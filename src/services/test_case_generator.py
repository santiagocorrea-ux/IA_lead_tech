"""
Generate VISA-format Excel test cases from a JSON definition file.

JSON schema expected:
    {
      "test_cases": [
        {
          "name": "Test case title",
          "steps": [
            {"action": "...", "data": "...", "expected_result": "..."}
          ]
        }
      ]
    }
"""

from __future__ import annotations

import json
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models.test_case import Step, TestCase


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def load_test_cases(json_path: str | Path) -> list[TestCase]:
    with open(json_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    raw_cases = payload.get("test_cases")
    if not isinstance(raw_cases, list) or not raw_cases:
        raise ValueError("JSON must contain a non-empty 'test_cases' array.")

    test_cases: list[TestCase] = []
    for i, raw_case in enumerate(raw_cases, start=1):
        if not isinstance(raw_case, dict):
            raise ValueError(f"test_cases[{i}] must be an object.")

        name = str(raw_case.get("name", "")).strip()
        raw_steps = raw_case.get("steps", [])

        if not name:
            raise ValueError(f"test_cases[{i}] is missing 'name'.")
        if not isinstance(raw_steps, list) or not raw_steps:
            raise ValueError(f"test_cases[{i}] must contain a non-empty 'steps' array.")

        steps: list[Step] = []
        for j, s in enumerate(raw_steps, start=1):
            if not isinstance(s, dict):
                raise ValueError(f"test_cases[{i}].steps[{j}] must be an object.")
            steps.append(Step(
                action=str(s.get("action", "") or ""),
                data=str(s.get("data", "") or ""),
                expected_result=str(s.get("expected_result", "") or ""),
            ))
        test_cases.append(TestCase(name=name, steps=steps))

    return test_cases


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _copy_cell_style(src_cell: Any, dst_cell: Any) -> None:
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = copy(src_cell.number_format)
    dst_cell.protection = copy(src_cell.protection)


def _copy_row_style(src_ws: Worksheet, src_row: int, dst_ws: Worksheet, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        _copy_cell_style(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col))


def _line_count(value: Any) -> int:
    text = "" if value is None else str(value)
    return max(1, text.count("\n") + 1) if text else 1


def _step_row_height(step: Step, base_height: float) -> float:
    max_lines = max(_line_count(step.action), _line_count(step.data), _line_count(step.expected_result))
    return max(base_height, 18 * max_lines + 4)


# ---------------------------------------------------------------------------
# Template reading
# ---------------------------------------------------------------------------

def _read_template_styles(template_path: str | Path) -> dict[str, Any]:
    """Extract style references and dimensions from the template's first sheet."""
    tmpl_wb = load_workbook(template_path)
    tmpl_ws = tmpl_wb[tmpl_wb.sheetnames[0]]

    title_row, header_row, detail_row, blank_row = 1, 2, 3, 6

    sample_heights = [
        tmpl_ws.row_dimensions[r].height
        for r in range(detail_row, blank_row)
        if tmpl_ws.row_dimensions[r].height
    ]

    return {
        "wb": tmpl_wb,
        "ws": tmpl_ws,
        "max_col": tmpl_ws.max_column,
        "title_row_idx": title_row,
        "header_row_idx": header_row,
        "detail_row_idx": detail_row,
        "blank_row_idx": blank_row,
        "name_label": tmpl_ws.cell(title_row, 1).value or "Name:",
        "headers": [
            tmpl_ws.cell(header_row, 1).value or "Action",
            tmpl_ws.cell(header_row, 2).value or "Data",
            tmpl_ws.cell(header_row, 3).value or "Expected Result",
        ],
        "title_height": tmpl_ws.row_dimensions[title_row].height or 24,
        "header_height": tmpl_ws.row_dimensions[header_row].height or 22,
        "detail_height": min(sample_heights) if sample_heights else 22,
        "blank_height": tmpl_ws.row_dimensions[blank_row].height or 10,
        "col_widths": {col: tmpl_ws.column_dimensions[col].width for col in tmpl_ws.column_dimensions},
    }


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def _write_cases_to_sheet(ws: Worksheet, test_cases: list[TestCase], styles: dict[str, Any]) -> None:
    tmpl_ws = styles["ws"]
    max_col = styles["max_col"]

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    for col_letter, width in styles["col_widths"].items():
        ws.column_dimensions[col_letter].width = width

    current_row = 1
    for case in test_cases:
        _copy_row_style(tmpl_ws, styles["title_row_idx"], ws, current_row, max_col)
        ws.row_dimensions[current_row].height = styles["title_height"]
        ws.cell(current_row, 1).value = styles["name_label"]
        ws.cell(current_row, 2).value = case.name
        current_row += 1

        _copy_row_style(tmpl_ws, styles["header_row_idx"], ws, current_row, max_col)
        ws.row_dimensions[current_row].height = styles["header_height"]
        for col, header in enumerate(styles["headers"], start=1):
            ws.cell(current_row, col).value = header
        current_row += 1

        for step in case.steps:
            _copy_row_style(tmpl_ws, styles["detail_row_idx"], ws, current_row, max_col)
            ws.row_dimensions[current_row].height = _step_row_height(step, styles["detail_height"])
            ws.cell(current_row, 1).value = step.action
            ws.cell(current_row, 2).value = step.data
            ws.cell(current_row, 3).value = step.expected_result
            current_row += 1

        _copy_row_style(tmpl_ws, styles["blank_row_idx"], ws, current_row, max_col)
        ws.row_dimensions[current_row].height = styles["blank_height"]
        for col in range(1, max_col + 1):
            ws.cell(current_row, col).value = None
        current_row += 1

    for row in range(current_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None
        ws.row_dimensions[row].height = None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_excel(
    template_path: str | Path,
    output_path: str | Path,
    test_cases: list[TestCase],
    sheet_name: str | None = None,
) -> Path:
    """
    Write test_cases into an Excel file at output_path.

    If output_path already exists, the target sheet is replaced in-place.
    If it does not exist, the template is used as the base workbook.

    Returns the resolved output path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    styles = _read_template_styles(template_path)

    if output_path.exists():
        wb = load_workbook(output_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
    else:
        wb = styles["wb"]
        ws = wb[wb.sheetnames[0]]
        if sheet_name and ws.title != sheet_name:
            ws.title = sheet_name

    _write_cases_to_sheet(ws, test_cases, styles)
    wb.save(output_path)
    return output_path
