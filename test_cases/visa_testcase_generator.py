#!/usr/bin/env python3
"""
Generate VISA-format manual test cases in Excel from a JSON file.

Usage:
    python visa_testcase_generator.py \
        --json sample_test_cases.json \
        --template VISASGF-4167_test_cases.xlsx \
        --output all_test_cases.xlsx

If --output already exists the script adds (or replaces) a sheet named after
the JSON file instead of creating a new workbook, so all user stories share
one Excel file.
"""

from __future__ import annotations

import argparse
import json
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, List
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class Step:
    action: str = ""
    data: str = ""
    expected_result: str = ""


@dataclass
class TestCase:
    name: str
    steps: List[Step]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create VISA-format Excel test cases from JSON."
    )
    parser.add_argument(
        "--json",
        required=True,
        help="Path to the input JSON file with test case definitions.",
    )
    parser.add_argument(
        "--template",
        required=True,
        help="Path to the VISA Excel template used as the style source.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path to the output Excel file (.xlsx). Existing files are updated in-place.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name to write into. Defaults to the JSON filename (without extension).",
    )
    return parser.parse_args()


def load_test_cases(json_path: str | Path) -> List[TestCase]:
    with open(json_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    raw_cases = payload.get("test_cases")
    if not isinstance(raw_cases, list) or not raw_cases:
        raise ValueError("JSON must contain a non-empty 'test_cases' array.")

    test_cases: List[TestCase] = []
    for index, raw_case in enumerate(raw_cases, start=1):
        if not isinstance(raw_case, dict):
            raise ValueError(f"test_cases[{index}] must be an object.")

        name = str(raw_case.get("name", "")).strip()
        raw_steps = raw_case.get("steps", [])

        if not name:
            raise ValueError(f"test_cases[{index}] is missing 'name'.")
        if not isinstance(raw_steps, list) or not raw_steps:
            raise ValueError(f"test_cases[{index}] must contain a non-empty 'steps' array.")

        steps: List[Step] = []
        for step_index, raw_step in enumerate(raw_steps, start=1):
            if not isinstance(raw_step, dict):
                raise ValueError(
                    f"test_cases[{index}].steps[{step_index}] must be an object."
                )

            steps.append(
                Step(
                    action=str(raw_step.get("action", "") or ""),
                    data=str(raw_step.get("data", "") or ""),
                    expected_result=str(raw_step.get("expected_result", "") or ""),
                )
            )

        test_cases.append(TestCase(name=name, steps=steps))

    return test_cases


def copy_cell_style(src_cell, dst_cell) -> None:
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = copy(src_cell.number_format)
    dst_cell.protection = copy(src_cell.protection)


def copy_template_row_style(src_ws: Worksheet, src_row: int, dst_ws: Worksheet, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        copy_cell_style(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col))


def clear_sheet_values(ws: Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None


def line_count(value: Any) -> int:
    text = "" if value is None else str(value)
    if not text:
        return 1
    return max(1, text.count("\n") + 1)


def step_row_height(step: Step, base_height: float) -> float:
    max_lines = max(
        line_count(step.action),
        line_count(step.data),
        line_count(step.expected_result),
    )
    # Keeps the original look for short rows, but grows when text wraps.
    return max(base_height, 18 * max_lines + 4)


def _read_template_styles(template_path: str | Path) -> dict:
    """Extract style references and dimensions from the template's first sheet."""
    tmpl_wb = load_workbook(template_path)
    tmpl_ws = tmpl_wb[tmpl_wb.sheetnames[0]]

    title_row_idx = 1
    header_row_idx = 2
    detail_row_idx = 3
    blank_row_idx = 6

    sample_detail_heights = [
        tmpl_ws.row_dimensions[r].height
        for r in range(detail_row_idx, blank_row_idx)
        if tmpl_ws.row_dimensions[r].height
    ]

    return {
        "wb": tmpl_wb,
        "ws": tmpl_ws,
        "max_col": tmpl_ws.max_column,
        "title_row_idx": title_row_idx,
        "header_row_idx": header_row_idx,
        "detail_row_idx": detail_row_idx,
        "blank_row_idx": blank_row_idx,
        "name_label": tmpl_ws.cell(title_row_idx, 1).value or "Name:",
        "headers": [
            tmpl_ws.cell(header_row_idx, 1).value or "Action",
            tmpl_ws.cell(header_row_idx, 2).value or "Data",
            tmpl_ws.cell(header_row_idx, 3).value or "Expected Result",
        ],
        "title_height": tmpl_ws.row_dimensions[title_row_idx].height or 24,
        "header_height": tmpl_ws.row_dimensions[header_row_idx].height or 22,
        "detail_height": min(sample_detail_heights) if sample_detail_heights else 22,
        "blank_height": tmpl_ws.row_dimensions[blank_row_idx].height or 10,
        "col_widths": {
            col: tmpl_ws.column_dimensions[col].width
            for col in tmpl_ws.column_dimensions
        },
    }


def _write_cases_to_sheet(ws: Worksheet, test_cases: List[TestCase], styles: dict) -> None:
    """Write test cases into ws, copying styles from the template sheet."""
    tmpl_ws = styles["ws"]
    max_col = styles["max_col"]

    clear_sheet_values(ws)

    # Copy column widths from template.
    for col_letter, width in styles["col_widths"].items():
        ws.column_dimensions[col_letter].width = width

    current_row = 1
    for case in test_cases:
        copy_template_row_style(tmpl_ws, styles["title_row_idx"], ws, current_row, max_col)
        ws.row_dimensions[current_row].height = styles["title_height"]
        ws.cell(current_row, 1).value = styles["name_label"]
        ws.cell(current_row, 2).value = case.name
        if max_col >= 3:
            ws.cell(current_row, 3).value = None
        current_row += 1

        copy_template_row_style(tmpl_ws, styles["header_row_idx"], ws, current_row, max_col)
        ws.row_dimensions[current_row].height = styles["header_height"]
        for col, header in enumerate(styles["headers"], start=1):
            ws.cell(current_row, col).value = header
        current_row += 1

        for step in case.steps:
            copy_template_row_style(tmpl_ws, styles["detail_row_idx"], ws, current_row, max_col)
            ws.row_dimensions[current_row].height = step_row_height(step, styles["detail_height"])
            ws.cell(current_row, 1).value = step.action
            ws.cell(current_row, 2).value = step.data
            ws.cell(current_row, 3).value = step.expected_result
            current_row += 1

        copy_template_row_style(tmpl_ws, styles["blank_row_idx"], ws, current_row, max_col)
        ws.row_dimensions[current_row].height = styles["blank_height"]
        for col in range(1, max_col + 1):
            ws.cell(current_row, col).value = None
        current_row += 1

    for row in range(current_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None
        ws.row_dimensions[row].height = None


def build_excel(
    template_path: str | Path,
    output_path: str | Path,
    test_cases: List[TestCase],
    sheet_name: str | None = None,
) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    styles = _read_template_styles(template_path)

    if output_path.exists():
        # Update existing workbook: replace the target sheet or add a new one.
        wb = load_workbook(output_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        # First run: initialise from template so sheet-level settings are preserved.
        wb = styles["wb"]
        ws = wb[wb.sheetnames[0]]
        if sheet_name and ws.title != sheet_name:
            ws.title = sheet_name

    _write_cases_to_sheet(ws, test_cases, styles)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    sheet_name = args.sheet or Path(args.json).stem
    test_cases = load_test_cases(args.json)
    output_path = Path(args.output)
    build_excel(
        template_path=args.template,
        output_path=output_path,
        test_cases=test_cases,
        sheet_name=sheet_name,
    )
    action = "Updated" if output_path.exists() else "Created"
    print(f"{action}: {output_path}  (sheet: '{sheet_name}')")


if __name__ == "__main__":
    main()
